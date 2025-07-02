import openpyxl
from lib.app.domain.entities import PartData

REQUIRED_COLS = ["part_number", "year", "category", "model", "vehicle_type", "color", "price", "stock"]

def parse_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = [cell.value.lower() for cell in next(sheet.iter_rows(max_row=1))]

    for col in REQUIRED_COLS:
        if col not in headers:
            raise ValueError(f"Missing required column: {col}")

    col_index = {col: headers.index(col) for col in REQUIRED_COLS}
    parts = []

    for row in sheet.iter_rows(min_row=2):
        data = {col: row[col_index[col]].value for col in REQUIRED_COLS}
        part = PartData(**data)
        parts.append(part)

    return parts
